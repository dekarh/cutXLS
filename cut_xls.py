# -*- coding: utf-8 -*-

import openpyxl
import sys
import time
import csv
from datetime import datetime
from lib import l, lenl, get_filename, get_path

IN_SNILS = ['СНИЛС', 'СтраховойНомер', 'Страховой_номер', 'Страховой Номер', 'Номер СНИЛС']


fields = []
workbooks =  []
sheets = []
for i, xlsx_file in enumerate(sys.argv):                              # Загружаем все xlsx файлы
    if i == 0:
        continue
    workbooks.append(openpyxl.load_workbook(filename=xlsx_file, read_only=True))
    sheets.append(workbooks[i-1][workbooks[i-1].sheetnames[0]])

sheets_keys = []
total_rows = 0
for i, sheet in enumerate(sheets):                                    # Маркируем нужные столбцы
    total_rows += sheet.max_row
    keys = {}
    for j, row in enumerate(sheet.rows):
        if j > 0:
            break
        for k, cell in enumerate(row):                                # Проверяем, чтобы был СНИЛС и Код
            if cell.value in IN_SNILS:
                keys[IN_SNILS[0]] = k
        if len(keys) < 1:
            print('В файле "' + sys.argv[i+1] + '" отсутствует колонка со СНИЛС')
            time.sleep(3)
            sys.exit()
    sheets_keys.append(keys)

path = get_path(sys.argv[1])

print('\n'+ datetime.now().strftime("%H:%M:%S") +' Начинаем преобразование и нарезку xlsx файлов \n')

k = 1                                                           # Счетчик строк в csv
file_number = 1
cl_csvs = []
for i, sheet in enumerate(sheets):                              # Загружаем все xlsx файлы по мере сохранения в БД
    for j, row in enumerate(sheet.rows):                              # Теперь строки
        if j == 0:
            continue
        cl_csv = {}
        if lenl(row[keys[IN_SNILS[0]]].value) < 12: # and l(row[keys[IN_SNILS[0]]].value) > 100:
            cl_csv[IN_SNILS[0]] = '{:=011d}'.format(l(row[keys[IN_SNILS[0]]].value))
            cl_csvs.append(cl_csv)
        else:
            print('\nСНИЛС ' + str(row[keys[IN_SNILS[0]]].value) + ' пропущен\n' )
        if k % 15000 == 0:
            with open(path + '{:=03d}'.format(file_number) + '.csv', 'w', encoding='cp1251') as output_file:
                dict_writer = csv.DictWriter(output_file, [IN_SNILS[0]], delimiter=';') #, quoting=csv.QUOTE_NONNUMERIC)
                dict_writer.writeheader()
                dict_writer.writerows(cl_csvs)
            output_file.close()
            cl_csvs = []
            k = 0
            print(datetime.now().strftime("%H:%M:%S") + ' 15k из файла '+ sys.argv[i+1].split(path)[1] +
                  ' сохранил в файл ' + '{:=03d}'.format(file_number) + '.csv')
            file_number += 1
        k += 1

    print('\n' + datetime.now().strftime("%H:%M:%S") + ' Файл '+ get_filename(sys.argv[i+1]) +' загружен полностью\n')
with open(path + '{:=03d}'.format(file_number) + '.csv', 'w', encoding='cp1251') as output_file:
    dict_writer = csv.DictWriter(output_file, [IN_SNILS[0]], delimiter=';')  # , quoting=csv.QUOTE_NONNUMERIC)
    dict_writer.writeheader()
    dict_writer.writerows(cl_csvs)
output_file.close()
print(datetime.now().strftime("%H:%M:%S") + ' Остаток из файла '+ get_filename(sys.argv[i+1]) +' сохранил в файл '
                  + '{:=03d}'.format(file_number) + '.csv')

print('\n'+ datetime.now().strftime("%H:%M:%S") +' Завершено без сбоев\n')